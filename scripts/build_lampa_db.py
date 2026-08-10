# -*- coding: utf-8 -*-
"""
lampa.db v2 = MySQL bulk (<5000) + νέο Excel μόνο για >=5000.
ΚΑΝΟΝΑΣ (>=5000): ονόματα υπαλλήλων & τμήματα/γραφεία που ΔΕΝ βρίσκονται σίγουρα
στη MySQL ΔΕΝ διαγράφονται — κρατιούνται στο *_original_text ΚΑΙ σημαίνονται
ως data_issue «προς εξέταση». Καμία αυτόματη μαντεψιά.
"""
import io
import os
import re
import sqlite3
import sys

import openpyxl
import pymysql

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

OUT = os.path.join(os.path.dirname(__file__), "lampa.db")
XLSX = r"E:\call_logger\Data Base\ΓΝΚ_Κατάλογος_Εξοπλισμού_από_παλιά_εφαρμογή_(Λάμπα).xlsx"

SCHEMA = [
    """CREATE TABLE offices (
        office INTEGER PRIMARY KEY, office_name TEXT, organization INTEGER,
        organization_name TEXT, department INTEGER, department_name TEXT,
        responsible INTEGER, responsible_original_text TEXT, e_mail TEXT,
        phones TEXT, building TEXT, level INTEGER)""",
    """CREATE TABLE owners (
        owner INTEGER PRIMARY KEY, last_name TEXT, first_name TEXT, office INTEGER,
        office_original_text TEXT, e_mail TEXT, phones TEXT,
        FOREIGN KEY (office) REFERENCES offices(office) ON DELETE RESTRICT ON UPDATE CASCADE)""",
    """CREATE TABLE model (
        model INTEGER PRIMARY KEY, model_name TEXT, category_code INTEGER,
        category_code_original_text TEXT, category_name TEXT, subcategory_code INTEGER,
        subcategory_code_original_text TEXT, subcategory_name TEXT, manufacturer INTEGER,
        manufacturer_original_text TEXT, manufacturer_name TEXT, manufacturer_code TEXT,
        attributes TEXT, consumables TEXT, network_connectivity INTEGER)""",
    """CREATE TABLE contracts (
        contract INTEGER PRIMARY KEY, contract_name TEXT, category INTEGER,
        category_original_text TEXT, category_name TEXT, supplier INTEGER,
        supplier_original_text TEXT, supplier_name TEXT, start_date TEXT, end_date TEXT,
        declaration TEXT, award TEXT, cost TEXT, committee TEXT, comments TEXT)""",
    """CREATE TABLE equipment (
        code INTEGER PRIMARY KEY, description TEXT, model INTEGER, model_original_text TEXT,
        serial_no TEXT, asset_no TEXT, state INTEGER, state_original_text TEXT, state_name TEXT,
        set_master INTEGER, set_master_original_text TEXT, contract INTEGER,
        contract_original_text TEXT, maintenance_contract TEXT, receiving_date TEXT,
        end_of_guarantee_date TEXT, cost TEXT, owner INTEGER, owner_original_text TEXT,
        office INTEGER, office_original_text TEXT, attributes TEXT, comments TEXT,
        ip_address TEXT, network_name TEXT, network_source TEXT, network_node TEXT,
        network_vlan TEXT, network_mac TEXT, network_description TEXT, network_comments TEXT,
        FOREIGN KEY (model) REFERENCES model(model) ON DELETE RESTRICT ON UPDATE CASCADE,
        FOREIGN KEY (contract) REFERENCES contracts(contract) ON DELETE SET NULL ON UPDATE CASCADE,
        FOREIGN KEY (owner) REFERENCES owners(owner) ON DELETE SET NULL ON UPDATE CASCADE,
        FOREIGN KEY (office) REFERENCES offices(office) ON DELETE RESTRICT ON UPDATE CASCADE,
        FOREIGN KEY (set_master) REFERENCES equipment(code) ON DELETE RESTRICT ON UPDATE CASCADE)""",
    """CREATE TABLE data_issues (
        id INTEGER PRIMARY KEY AUTOINCREMENT, sheet TEXT, row_number INTEGER,
        column_name TEXT, raw_value TEXT, issue_type TEXT NOT NULL, message TEXT,
        created_at TEXT NOT NULL, status TEXT NOT NULL DEFAULT 'open')""",
]


def s(v):
    if v is None:
        return None
    if isinstance(v, str):
        return v.strip() or None
    return str(v)


def is_int(v):
    try:
        int(v); return True
    except (TypeError, ValueError):
        return False


if not os.path.exists(XLSX):
    print("✗ ΔΕΝ ΥΠΑΡΧΕΙ το Excel:", XLSX); sys.exit(2)

# ------------------------------------------------------------ MySQL bulk
my = pymysql.connect(host="192.168.12.24", user="lampa_ro", password="1234",
                     database="equipment", charset="utf8",
                     cursorclass=pymysql.cursors.DictCursor)
mc = my.cursor()

if os.path.exists(OUT):
    os.remove(OUT)
db = sqlite3.connect(OUT)
db.execute("PRAGMA foreign_keys=OFF")
for stmt in SCHEMA:
    db.execute(stmt)

mc.execute("""SELECT o.code office, o.name office_name, d.organization organization,
    org.name organization_name, o.department department, d.name department_name,
    o.responsible, o.e_mail, o.phones, o.building, o.level
    FROM offices o LEFT JOIN departments d ON d.code=o.department
    LEFT JOIN organizations org ON org.code=d.organization""")
for r in mc.fetchall():
    db.execute("""INSERT INTO offices(office,office_name,organization,organization_name,
        department,department_name,responsible,responsible_original_text,e_mail,phones,building,level)
        VALUES(?,?,?,?,?,?,?,NULL,?,?,?,?)""",
        (r["office"], s(r["office_name"]), r["organization"], s(r["organization_name"]),
         r["department"], s(r["department_name"]), r["responsible"], s(r["e_mail"]),
         s(r["phones"]), s(r["building"]), r["level"]))

mc.execute("SELECT code owner,last_name,first_name,office,e_mail,phones FROM employees")
for r in mc.fetchall():
    db.execute("""INSERT INTO owners(owner,last_name,first_name,office,office_original_text,e_mail,phones)
        VALUES(?,?,?,?,NULL,?,?)""",
        (r["owner"], s(r["last_name"]), s(r["first_name"]), r["office"], s(r["e_mail"]), s(r["phones"])))

mc.execute("""SELECT m.code model, m.name model_name, sc.category category_code,
    cat.name category_name, m.subcategory subcategory_code, sc.name subcategory_name,
    m.manufacturer, tr.name manufacturer_name, m.manufacturer_code, m.attributes,
    m.consumables, m.network_connectivity
    FROM models m LEFT JOIN equipment_subcategories sc ON sc.code=m.subcategory
    LEFT JOIN equipment_categories cat ON cat.code=sc.category
    LEFT JOIN traders tr ON tr.code=m.manufacturer""")
model_name_to_code = {}
max_model = 0
for r in mc.fetchall():
    db.execute("""INSERT INTO model(model,model_name,category_code,category_code_original_text,
        category_name,subcategory_code,subcategory_code_original_text,subcategory_name,
        manufacturer,manufacturer_original_text,manufacturer_name,manufacturer_code,
        attributes,consumables,network_connectivity)
        VALUES(?,?,?,NULL,?,?,NULL,?,?,NULL,?,?,?,?,?)""",
        (r["model"], s(r["model_name"]), r["category_code"], s(r["category_name"]),
         r["subcategory_code"], s(r["subcategory_name"]), r["manufacturer"],
         s(r["manufacturer_name"]), s(r["manufacturer_code"]), s(r["attributes"]),
         s(r["consumables"]), r["network_connectivity"]))
    if r["model_name"]:
        model_name_to_code.setdefault(r["model_name"].strip().lower(), r["model"])
    max_model = max(max_model, r["model"])

mc.execute("""SELECT c.code contract, c.contract contract_name, c.category,
    cc.name category_name, c.supplier, tr.name supplier_name, c.start_date, c.end_date,
    c.declaration, c.award, c.cost, c.committee, c.comments
    FROM contracts c LEFT JOIN contract_categories cc ON cc.code=c.category
    LEFT JOIN traders tr ON tr.code=c.supplier""")
valid_contracts = set()
for r in mc.fetchall():
    db.execute("""INSERT INTO contracts(contract,contract_name,category,category_original_text,
        category_name,supplier,supplier_original_text,supplier_name,start_date,end_date,
        declaration,award,cost,committee,comments)
        VALUES(?,?,?,NULL,?,?,NULL,?,?,?,?,?,?,?,?)""",
        (r["contract"], s(r["contract_name"]), r["category"], s(r["category_name"]),
         r["supplier"], s(r["supplier_name"]), s(r["start_date"]), s(r["end_date"]),
         s(r["declaration"]), s(r["award"]), s(r["cost"]), s(r["committee"]), s(r["comments"])))
    valid_contracts.add(r["contract"])

mc.execute("SELECT code,name FROM equipment_states")
state_name = {r["code"]: r["name"] for r in mc.fetchall()}

mc.execute("""SELECT equipment, MIN(code) node, ip1, ip2, ip3, ip4, mac, vlan,
    network_name nn, workgroup wg, comments cm
    FROM network_nodes WHERE equipment IS NOT NULL GROUP BY equipment""")
net_by_eq = {}
for r in mc.fetchall():
    ip = None
    if all(r[k] is not None for k in ("ip1", "ip2", "ip3", "ip4")):
        ip = f"{r['ip1']}.{r['ip2']}.{r['ip3']}.{r['ip4']}"
    net_by_eq[r["equipment"]] = dict(ip=ip, mac=s(r["mac"]), vlan=s(r["vlan"]),
        name=s(r["nn"]), node=str(r["node"]), wg=s(r["wg"]), cm=s(r["cm"]))

selfmaster = 0
mc.execute("""SELECT code,description,model,serial_no,asset_no,state,set_master,contract,
    maintenance_contract,receiving_date,end_of_guarantee_date,cost,owner,office,attributes,comments
    FROM equipment""")
for r in mc.fetchall():
    sm = r["set_master"]
    if sm is not None and sm == r["code"]:
        sm = None; selfmaster += 1
    net = net_by_eq.get(r["code"], {})
    db.execute("""INSERT INTO equipment(code,description,model,model_original_text,serial_no,
        asset_no,state,state_original_text,state_name,set_master,set_master_original_text,
        contract,contract_original_text,maintenance_contract,receiving_date,end_of_guarantee_date,
        cost,owner,owner_original_text,office,office_original_text,attributes,comments,
        ip_address,network_name,network_source,network_node,network_vlan,network_mac,
        network_description,network_comments)
        VALUES(?,?,?,NULL,?,?,?,NULL,?,?,NULL,?,NULL,?,?,?,?,?,NULL,?,NULL,?,?,?,?,?,?,?,?,?,?)""",
        (r["code"], s(r["description"]), r["model"], s(r["serial_no"]), s(r["asset_no"]),
         r["state"], s(state_name.get(r["state"])), sm, r["contract"],
         s(r["maintenance_contract"]), s(r["receiving_date"]), s(r["end_of_guarantee_date"]),
         s(r["cost"]), r["owner"], r["office"], s(r["attributes"]), s(r["comments"]),
         net.get("ip"), net.get("name"), "lampa" if net else None, net.get("node"),
         net.get("vlan"), net.get("mac"), net.get("wg"), net.get("cm")))
my.close()

# ------------------------------------------------------------ maps για ΣΥΝΤΗΡΗΤΙΚΟ ταίριασμα
_ACC = str.maketrans("άέήίόύώϊϋΐΰςΆΈΉΊΌΎΏ", "αεηιουωιυιυσαεηιουω")
def gnorm(x):
    t = (str(x) if x is not None else "").translate(_ACC).casefold()
    return " ".join(t.split())

existing_owners = {row[0] for row in db.execute("SELECT owner FROM owners")}
existing_offices = {row[0] for row in db.execute("SELECT office FROM offices")}
owner_full_to_ids, owner_last_to_ids = {}, {}
for oid, ln, fn in db.execute("SELECT owner,last_name,first_name FROM owners"):
    lnn, fnn = gnorm(ln), gnorm(fn)
    for key in {f"{lnn} {fnn}".strip(), f"{fnn} {lnn}".strip()}:
        if key:
            owner_full_to_ids.setdefault(key, set()).add(oid)
    if lnn:
        owner_last_to_ids.setdefault(lnn, set()).add(oid)
office_name_to_ids = {}
for ofc, onm in db.execute("SELECT office,office_name FROM offices"):
    k = gnorm(onm)
    if k:
        office_name_to_ids.setdefault(k, set()).add(ofc)


def match_owner(val):
    if val is None or str(val).strip() == "":
        return None, None, None
    if is_int(val):
        oid = int(val)
        return (oid, None, None) if oid in existing_owners else (None, str(oid), "Άγνωστος κωδικός υπαλλήλου — προς εξέταση.")
    key = gnorm(val)
    ids = owner_full_to_ids.get(key)
    if ids and len(ids) == 1:
        return next(iter(ids)), None, None
    # Option B: μοναδικό επώνυμο σε οποιοδήποτε token → σύνδεση (καμία ασάφεια)
    cand = set()
    for tok in key.split():
        sids = owner_last_to_ids.get(tok)
        if sids and len(sids) == 1:
            cand |= sids
    if len(cand) == 1:
        return next(iter(cand)), None, None
    return None, str(val).strip(), "Όνομα υπαλλήλου χωρίς σίγουρη αντιστοίχιση — προς εξέταση."


def match_office(val):
    if val is None or str(val).strip() == "":
        return None, None, None
    if is_int(val):
        ofid = int(val)
        return (ofid, None, None) if ofid in existing_offices else (None, str(ofid), "Άγνωστος κωδικός γραφείου — προς εξέταση.")
    ids = office_name_to_ids.get(gnorm(val))
    if ids and len(ids) == 1:
        return next(iter(ids)), None, None
    return None, str(val).strip(), "Τμήμα/γραφείο χωρίς σίγουρη αντιστοίχιση — προς εξέταση."


# ------------------------------------------------------------ Excel (>=5000) με βάση επικεφαλίδες
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
def header_map(ws):
    hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or ()
    return {str(h).strip().lower(): i for i, h in enumerate(hdr) if h is not None}

eq_ws, eq_h = None, {}
for sh in wb.sheetnames:
    hm = header_map(wb[sh])
    if "code" in hm and ("owner" in hm or "office" in hm):
        eq_ws, eq_h = wb[sh], hm; break
if eq_ws is None:
    print("✗ Δεν βρέθηκε φύλλο εξοπλισμού. Φύλλα/στήλες:")
    for sh in wb.sheetnames:
        print(f"  «{sh}»: {list(header_map(wb[sh]).keys())}")
    sys.exit(3)
print(f"[Excel] φύλλο='{eq_ws.title}'  στήλες={list(eq_h.keys())}")


def col(row, *names):
    for n in names:
        if n in eq_h:
            return row[eq_h[n]]
    return None

ge = [row for row in eq_ws.iter_rows(min_row=2, values_only=True)
      if is_int(col(row, "code")) and int(col(row, "code")) >= 5000]
wb.close()
print(f"[Excel] γραμμές code>=5000: {len(ge)}")

MAC_RE = re.compile(r"([0-9A-Fa-f]{2}(?:[:\-][0-9A-Fa-f]{2}){5})|(?<![0-9A-Fa-f])([0-9A-Fa-f]{12})(?![0-9A-Fa-f])")
PORT_RE = re.compile(r"θυρα\s*(\d+)", re.IGNORECASE)
def norm_mac(m):
    hexs = re.sub(r"[^0-9A-Fa-f]", "", m).upper()
    return ":".join(hexs[i:i+2] for i in range(0, 12, 2)) if len(hexs) == 12 else m

new_model_code = max_model
issues = []
now = "2026-07-22T00:00:00"
for row in ge:
    code = int(col(row, "code"))
    desc = s(col(row, "description"))
    mdl_txt = s(col(row, "model", "model name"))
    serial = s(col(row, "serial_no", "serial no"))
    asset = s(col(row, "asset_no", "asset no"))
    st_v = col(row, "state")
    state = int(st_v) if is_int(st_v) else None
    contract_raw = col(row, "contract")
    owner_raw = col(row, "owner", "owner name")
    office_raw = col(row, "office", "office name", "department", "department name")
    comments = s(col(row, "comments"))

    model_id = model_orig = None
    if mdl_txt:
        key = mdl_txt.strip().lower()
        if key in model_name_to_code:
            model_id = model_name_to_code[key]
        else:
            new_model_code += 1
            db.execute("INSERT INTO model(model,model_name) VALUES(?,?)", (new_model_code, mdl_txt))
            model_name_to_code[key] = new_model_code
            model_id = new_model_code

    owner_id, owner_orig, owner_issue = match_owner(owner_raw)
    if owner_issue:
        issues.append((code, "owner", owner_orig or str(owner_raw), owner_issue))
    office_id, office_orig, office_issue = match_office(office_raw)
    if office_issue:
        issues.append((code, "office", office_orig or str(office_raw), office_issue))

    contract_id = contract_orig = None
    if is_int(contract_raw) and int(contract_raw) in valid_contracts:
        contract_id = int(contract_raw)
    elif contract_raw not in (None, ""):
        contract_orig = str(contract_raw).strip()

    ip = nname = nsrc = nnode = nvlan = nmac = ndesc = ncom = None
    if comments:
        mm = MAC_RE.search(comments); pm = PORT_RE.search(comments)
        if mm:
            nmac = norm_mac(mm.group(0)); nsrc = "excel-comments"; ncom = comments
        if pm:
            ndesc = f"Θύρα {pm.group(1)}"; nsrc = "excel-comments"; ncom = comments

    db.execute("""INSERT INTO equipment(code,description,model,model_original_text,serial_no,asset_no,
        state,state_original_text,state_name,set_master,set_master_original_text,contract,
        contract_original_text,maintenance_contract,receiving_date,end_of_guarantee_date,cost,
        owner,owner_original_text,office,office_original_text,attributes,comments,
        ip_address,network_name,network_source,network_node,network_vlan,network_mac,
        network_description,network_comments)
        VALUES(?,?,?,?,?,?,?,NULL,?,NULL,NULL,?,?,NULL,NULL,NULL,NULL,?,?,?,?,NULL,?,?,?,?,?,?,?,?,?)""",
        (code, desc, model_id, model_orig, serial, asset, state, s(state_name.get(state)),
         contract_id, contract_orig, owner_id, owner_orig, office_id, office_orig, comments,
         ip, nname, nsrc, nnode, nvlan, nmac, ndesc, ncom))

for (cd, coln, raw, msg) in issues:
    db.execute("""INSERT INTO data_issues(sheet,row_number,column_name,raw_value,issue_type,message,created_at,status)
        VALUES('equipment',?,?,?,'unknown_id',?,?,'open')""", (cd, coln, str(raw), msg, now))
db.commit()

# ------------------------------------------------------------ ΕΠΙΚΥΡΩΣΗ
def q(sql):
    return db.execute(sql).fetchone()[0]

print("\n" + "=" * 60)
print("lampa.db (v2) χτίστηκε:", OUT)
print("=" * 60)
for t in ["offices", "owners", "model", "contracts", "equipment", "data_issues"]:
    print(f"  {t:12s}: {q(f'SELECT COUNT(*) FROM {t}')}")
print("\n--- ΕΞΟΠΛΙΣΜΟΣ ---")
print("  MySQL (<5000):", q("SELECT COUNT(*) FROM equipment WHERE code<5000"),
      "| Excel (>=5000):", q("SELECT COUNT(*) FROM equipment WHERE code>=5000"))
print("  νέα μοντέλα:", new_model_code - max_model, "| self-master μηδενισμένα:", selfmaster)

print("\n--- ΟΡΦΑΝΑ FK (πρέπει ~0) ---")
for c, t, k in [("model", "model", "model"), ("owner", "owners", "owner"),
                ("office", "offices", "office"), ("contract", "contracts", "contract")]:
    print(f"  {c}:", q(f"SELECT COUNT(*) FROM equipment e WHERE {c} IS NOT NULL AND NOT EXISTS(SELECT 1 FROM {t} x WHERE x.{k}=e.{c})"))

print("\n--- «ΠΡΟΣ ΕΞΕΤΑΣΗ» (>=5000) ---")
print("  σύνολο data_issues:", q("SELECT COUNT(*) FROM data_issues"))
print("  owner προς εξέταση:", q("SELECT COUNT(*) FROM data_issues WHERE column_name='owner'"))
print("  office/τμήμα προς εξέταση:", q("SELECT COUNT(*) FROM data_issues WHERE column_name='office'"))
print("  δείγμα:")
for r in db.execute("SELECT row_number,column_name,raw_value,message FROM data_issues ORDER BY row_number LIMIT 40"):
    print(f"    code {r[0]} | {r[1]} = '{r[2]}' | {r[3]}")

print("\n--- ΔΙΑΤΗΡΗΘΗΚΑΝ ΟΝΟΜΑΤΑ/ΤΜΗΜΑΤΑ ΑΥΤΟΥΣΙΑ (δείγμα) ---")
for r in db.execute("""SELECT code,owner_original_text,office_original_text FROM equipment
    WHERE code>=5000 AND (owner_original_text IS NOT NULL OR office_original_text IS NOT NULL) LIMIT 20"""):
    print(f"    code {r[0]} | owner_txt={r[1]!r} | office_txt={r[2]!r}")

db.close()
